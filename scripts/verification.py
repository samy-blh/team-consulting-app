import time
import sys
import pandas as pd
import pytz
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from pathlib import Path
import unicodedata

tz_paris = pytz.timezone("Europe/Paris")
fichier_excel = sys.argv[1]
fichier_sortie = sys.argv[2]

options = Options()
options.add_argument("--headless")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")
options.add_argument("--window-size=1280,1024")

df = pd.read_excel(fichier_excel)
interventions_a_suivre = []

def extraire_interventions(driver, nom, login, onglet_type):
    try:
        driver.find_element(By.LINK_TEXT, onglet_type).click()
        time.sleep(4)
        cards = driver.find_elements(By.CLASS_NAME, "intervention")
        for i in range(len(cards)):
            try:
                cards = driver.find_elements(By.CLASS_NAME, "intervention")
                card = cards[i]
                card.click()
                time.sleep(2)

                rdv_time = None
                jeton_val = ""
                adresse_client = ""
                debut_intervention = ""

                labels = driver.find_elements(By.CLASS_NAME, "label")
                for label in labels:
                    try:
                        b = label.find_element(By.TAG_NAME, "b")
                        label_title = b.text.strip().lower()
                        texte_complet = label.text.strip()

                        if "date du rdv" in label_title:
                            date_str = texte_complet.split(":")[1].strip()
                            if len(date_str) == 13:
                                date_str += ":00"
                            rdv_time = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
                            rdv_time = tz_paris.localize(rdv_time)

                        elif "jeton" in label_title:
                            jeton_val = texte_complet.split(":")[1].strip()

                        elif "adresse" in label_title:
                            try:
                                adresse_client = label.find_element(By.TAG_NAME, "a").text.strip()
                            except:
                                adresse_client = texte_complet.split(":")[1].strip()

                        elif "début" in label_title:
                            debut_intervention = texte_complet.split(":")[1].strip()
                    except:
                        continue

                now = datetime.now(tz_paris)

                if not rdv_time or now.date() != rdv_time.date():
                    driver.back()
                    time.sleep(1)
                    continue

                if debut_intervention:
                    statut = f"Démarrée à {debut_intervention}"
                else:
                    if now > rdv_time + timedelta(minutes=10):
                        statut = "Non démarrée - En retard"
                    else:
                        statut = "À venir - Non démarrée"

                interventions_a_suivre.append({
                    "technicien": nom,
                    "login": login,
                    "jeton": jeton_val,
                    "adresse": adresse_client,
                    "rdv": rdv_time.strftime("%Y-%m-%d %H:%M"),
                    "statut": statut,
                    "heure_actuelle": now.strftime("%Y-%m-%d %H:%M"),
                    "type": onglet_type
                })

                driver.back()
                time.sleep(1)

            except Exception as e:
                print(f"Erreur sur une intervention : {e}")
                continue
    except Exception as e:
        print(f"Erreur onglet {onglet_type} : {e}")

for index, row in df.iterrows():
    nom = row["nom"]
    login = str(row["login"])
    password = str(row["password"])

    print(f"Connexion pour {nom}...")

    driver = webdriver.Chrome(options=options)
    driver.get("https://aboracco.pub.app.ftth.iliad.fr/")
    time.sleep(3)

    inputs = driver.find_elements(By.TAG_NAME, "input")
    inputs[0].send_keys(login)
    inputs[1].send_keys(password)
    driver.find_element(By.XPATH, "//button[contains(text(), 'Connexion')]").click()
    time.sleep(4)

    extraire_interventions(driver, nom, login, "Production")
    extraire_interventions(driver, nom, login, "Post-Production / SAV")

    driver.quit()

if interventions_a_suivre:
    Path(fichier_sortie).parent.mkdir(parents=True, exist_ok=True)
    df_result = pd.DataFrame(interventions_a_suivre)
    df_result.to_excel(fichier_sortie, index=False)

    wb = load_workbook(fichier_sortie)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    fill_orange = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    statut_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "statut":
            statut_col = idx

    def normalize(s):
        if not s:
            return ""
        return unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('utf-8').lower()

    if statut_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            statut_cell = row[statut_col - 1]
            apply_fill = False
            if "non demarree - en retard" in normalize(statut_cell.value):
                apply_fill = True
            for cell in row:
                cell.border = border
                if apply_fill:
                    cell.fill = fill_orange

    wb.save(fichier_sortie)
