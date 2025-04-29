import time
import pandas as pd
import sys
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import unicodedata
import subprocess

# Demander le fichier Excel et la date
fichier = input("Entrez le nom du fichier Excel (ex: liste_techniciens.xlsx) : ")
try:
    df = pd.read_excel(fichier)
except Exception as e:
    print(f"Erreur lors de la lecture du fichier : {e}")
    sys.exit()

date_input = input("Entrez la date des interventions au format JJ/MM/AAAA : ")
try:
    date_cible = datetime.strptime(date_input, "%d/%m/%Y").date()
except ValueError:
    print("Format de date invalide. Utilisez JJ/MM/AAAA.")
    sys.exit()

options = Options()
options.add_argument("--start-maximized")

interventions_terminees = []

def extraire_interventions_terminees(driver, nom, login, onglet_type):
    try:
        driver.find_element(By.LINK_TEXT, onglet_type).click()
        time.sleep(4)

        boutons = driver.find_elements(By.CLASS_NAME, "btn-outline-danger")
        for btn in boutons:
            if "Terminées" in btn.text:
                btn.click()
                time.sleep(4)
                break

        cards = driver.find_elements(By.CLASS_NAME, "intervention")
        for i in range(len(cards)):
            try:
                cards = driver.find_elements(By.CLASS_NAME, "intervention")
                card = cards[i]
                text = card.text
                lines = text.split("\n")
                date_line = next((l for l in lines if "Date du RDV" in l), None)
                if not date_line:
                    continue
                date_str = date_line.split(":")[1].strip()
                if len(date_str) == 13:
                    date_str += ":00"
                rdv_time = datetime.strptime(date_str, "%Y-%m-%d %H:%M")

                if rdv_time.date() != date_cible:
                    continue

                card.click()
                time.sleep(2)

                debut_intervention = ""
                fin_intervention = ""
                jeton_val = ""
                adresse_client = ""
                etat_box = "Non défini"

                labels = driver.find_elements(By.CLASS_NAME, "label")
                for label in labels:
                    try:
                        b = label.find_element(By.TAG_NAME, "b")
                        label_title = b.text.strip().lower()
                        texte_complet = label.text.strip()

                        if "début de l'intervention" in label_title:
                            parts = texte_complet.split(":")
                            if len(parts) > 1:
                                debut_intervention = parts[1].strip()

                        elif "fin de l'intervention" in label_title:
                            parts = texte_complet.split(":")
                            if len(parts) > 1:
                                fin_intervention = parts[1].strip()

                        elif "jeton" in label_title:
                            parts = texte_complet.split(":")
                            if len(parts) > 1:
                                jeton_val = parts[1].strip()

                        elif "adresse" in label_title:
                            try:
                                adresse_client = label.find_element(By.TAG_NAME, "a").text.strip()
                            except:
                                adresse_client = texte_complet.split(":")[1].strip()
                    except:
                        continue

                try:
                    etats_divs = driver.find_elements(By.XPATH, "//div[@style]")
                    for div in etats_divs:
                        texte = div.text.strip()
                        if texte in ["OK", "NOK"]:
                            etat_box = texte
                            break
                except:
                    pass

                interventions_terminees.append({
                    "technicien": nom,
                    "login": login,
                    "jeton": jeton_val,
                    "adresse": adresse_client,
                    "rdv": rdv_time.strftime("%Y-%m-%d %H:%M"),
                    "début": debut_intervention,
                    "fin": fin_intervention,
                    "heure_actuelle": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "etat_box": etat_box,
                    "type": onglet_type
                })

                driver.back()
                time.sleep(2)

            except Exception as e:
                print(f"Erreur sur une intervention : {e}")
                continue

    except Exception as e:
        print(f"Erreur pour {nom} dans l’onglet {onglet_type} : {e}")

for index, row in df.iterrows():
    nom = row["nom"]
    login = str(row["login"])
    password = str(row["password"])

    print(f"Connexion pour {nom}...")

    driver = webdriver.Chrome(options=options)
    driver.get("https://aboracco.pub.app.ftth.iliad.fr/")
    time.sleep(3)

    inputs = driver.find_elements(By.TAG_NAME, "input")
    inputs[0].send_keys(login)
    inputs[1].send_keys(password)

    time.sleep(1)
    bouton_connexion = driver.find_element(By.XPATH, "//button[contains(text(), 'Connexion')]")
    bouton_connexion.click()
    time.sleep(4)

    extraire_interventions_terminees(driver, nom, login, "Production")
    extraire_interventions_terminees(driver, nom, login, "Post-Production / SAV")

    driver.quit()

if interventions_terminees:
    now = datetime.now()
    fichier_date = now.strftime("%d-%m-%Y %H-%M")
    nom_fichier = f"etat_des_terminees_{fichier_date}.xlsx"

    df_result = pd.DataFrame(interventions_terminees)
    df_result.to_excel(nom_fichier, index=False)

    wb = load_workbook(nom_fichier)
    ws = wb.active

    # Ajustement automatique des colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Bordures et couleur NOK
    fill_orange = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    etat_box_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "etat_box":
            etat_box_col = idx

    if etat_box_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            etat_cell = row[etat_box_col - 1]
            apply_fill = False
            if str(etat_cell.value).strip().upper() == "NOK":
                apply_fill = True
            for cell in row:
                cell.border = border
                if apply_fill:
                    cell.fill = fill_orange

    wb.save(nom_fichier)
    subprocess.Popen(["start", "", nom_fichier], shell=True)
    print(f"\nFichier '{nom_fichier}' généré avec succès.")
else:
    print("\nAucune intervention terminée détectée pour la date spécifiée.")
